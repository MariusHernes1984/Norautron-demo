from __future__ import annotations

import hashlib
import json
import logging
import math
import os
import re
import struct
import sys
import time
import uuid
from contextlib import contextmanager
from collections.abc import Callable, Iterable, Iterator, Sequence
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse

import pyodbc
from azure.core import MatchConditions
from azure.identity import DefaultAzureCredential, ManagedIdentityCredential
from azure.storage.blob import BlobClient
from openpyxl import load_workbook

MODULE_DIR = Path(__file__).resolve().parent
SCHEMA_PATH = next(
    (
        candidate
        for candidate in (
            MODULE_DIR / "data" / "schema.json",
            MODULE_DIR.parent / "data" / "schema.json",
        )
        if candidate.is_file()
    ),
    MODULE_DIR / "data" / "schema.json",
)
ACCESS_TOKEN_ATTRIBUTE = 1256
SQL_SCOPE = "https://database.windows.net/.default"
IDENTIFIER = re.compile(r"^[a-z][a-z0-9_]*$")
SUPPORTED_TYPES = {"string", "integer", "decimal", "date", "datetime"}
BATCH_SIZE = 1000
LOGGER_NAME = "norautron.ingest"
PRIVATE_LOG_FIELD = re.compile(
    r"(^|_)(address|chat|content|ip|message|prompt|question|sql|text)(_|$)",
    re.IGNORECASE,
)


def required(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"{name} is required")
    return value


def credential() -> DefaultAzureCredential | ManagedIdentityCredential:
    if os.getenv("ENVIRONMENT", "production").lower() == "development":
        return DefaultAzureCredential()
    client_id = os.getenv("AZURE_CLIENT_ID")
    return (
        ManagedIdentityCredential(client_id=client_id)
        if client_id
        else ManagedIdentityCredential()
    )


logging.basicConfig(level=logging.INFO, format="%(message)s")
LOGGER = logging.getLogger(LOGGER_NAME)
_observability_error: str | None = None
if os.getenv("APPLICATIONINSIGHTS_CONNECTION_STRING"):
    try:
        from azure.monitor.opentelemetry import configure_azure_monitor

        configure_azure_monitor(
            connection_string=os.environ["APPLICATIONINSIGHTS_CONNECTION_STRING"],
            credential=credential(),
            logger_name=LOGGER_NAME,
            enable_live_metrics=False,
        )
    except Exception as error:
        _observability_error = type(error).__name__

try:
    from opentelemetry import metrics

    _meter = metrics.get_meter("norautron.ingest", "0.1.0")
    _stage_duration = _meter.create_histogram(
        "norautron.ingest.stage.duration", unit="ms"
    )
    _run_count = _meter.create_counter("norautron.ingest.run.count")
except ImportError:
    _stage_duration = None
    _run_count = None


def log_event(
    event: str,
    *,
    level: int = logging.INFO,
    **attributes: str | int | float | bool | None,
) -> None:
    safe_attributes = {
        key: "[redacted]" if PRIVATE_LOG_FIELD.search(key) else value
        for key, value in attributes.items()
        if value is not None
    }
    LOGGER.log(
        level,
        json.dumps(
            {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "severity": logging.getLevelName(level).lower(),
                "event": event,
                "service": "norautron-ingest",
                **safe_attributes,
            },
            ensure_ascii=False,
            sort_keys=True,
        ),
    )


@contextmanager
def observed_stage(
    stage: str, **attributes: str | int | float | bool | None
) -> Iterator[None]:
    started = time.perf_counter()
    log_event("etl_stage_started", stage=stage, **attributes)
    try:
        yield
    except Exception as error:
        duration_ms = (time.perf_counter() - started) * 1000
        if _stage_duration is not None:
            _stage_duration.record(
                duration_ms, {"stage": stage, "outcome": "failure"}
            )
        log_event(
            "etl_stage_completed",
            level=logging.ERROR,
            stage=stage,
            outcome="failure",
            duration_ms=round(duration_ms),
            error_type=type(error).__name__,
            **attributes,
        )
        raise
    else:
        duration_ms = (time.perf_counter() - started) * 1000
        if _stage_duration is not None:
            _stage_duration.record(
                duration_ms, {"stage": stage, "outcome": "success"}
            )
        log_event(
            "etl_stage_completed",
            stage=stage,
            outcome="success",
            duration_ms=round(duration_ms),
            **attributes,
        )


if _observability_error:
    log_event(
        "telemetry_initialization_failed",
        level=logging.ERROR,
        error_type=_observability_error,
    )
elif os.getenv("APPLICATIONINSIGHTS_CONNECTION_STRING"):
    log_event("telemetry_initialized")


def pack_access_token(token: str) -> bytes:
    token_bytes = token.encode("utf-16-le")
    return struct.pack(f"<I{len(token_bytes)}s", len(token_bytes), token_bytes)


def sql_connection() -> pyodbc.Connection:
    token = credential().get_token(SQL_SCOPE).token
    connection_string = (
        "Driver={ODBC Driver 18 for SQL Server};"
        f"Server=tcp:{required('SQL_SERVER')},1433;"
        f"Database={required('SQL_DATABASE')};"
        "Encrypt=yes;TrustServerCertificate=no;Connection Timeout=30;"
    )
    return pyodbc.connect(
        connection_string,
        attrs_before={ACCESS_TOKEN_ATTRIBUTE: pack_access_token(token)},
        autocommit=False,
    )


def blob_name(blob_url: str) -> str:
    parsed = urlparse(blob_url)
    name = unquote(parsed.path.rsplit("/", 1)[-1])
    if not parsed.scheme == "https" or not parsed.netloc or not name:
        raise ValueError("DATA_BLOB_URL must be a complete HTTPS blob URL")
    return name


def download_workbook(
    target: Path,
    *,
    blob_factory: Callable[..., Any] = BlobClient.from_blob_url,
    token_credential: Any | None = None,
) -> tuple[str, str, int]:
    url = required("DATA_BLOB_URL")
    blob = blob_factory(url, credential=token_credential or credential())
    properties = blob.get_blob_properties()
    etag = properties.etag
    if not etag:
        raise RuntimeError("Source blob did not provide an ETag")
    digest = hashlib.sha256()
    target.parent.mkdir(parents=True, exist_ok=True)
    downloader = blob.download_blob(
        max_concurrency=4,
        etag=etag,
        match_condition=MatchConditions.IfNotModified,
    )
    with target.open("wb") as handle:
        for chunk in downloader.chunks():
            handle.write(chunk)
            digest.update(chunk)
    return str(etag), digest.hexdigest(), target.stat().st_size


def load_schema(path: Path = SCHEMA_PATH) -> dict[str, Any]:
    schema = json.loads(path.read_text(encoding="utf-8"))
    validate_schema(schema)
    return schema


def validate_schema(schema: dict[str, Any]) -> None:
    sources = schema.get("sources")
    if not isinstance(sources, list) or len(sources) != 5:
        raise ValueError("Schema must define exactly five sources")
    expected_total = schema.get("expectedTotalRows")
    if not isinstance(expected_total, int) or expected_total <= 0:
        raise ValueError("Schema expectedTotalRows must be a positive integer")

    tables: set[str] = set()
    sheets: set[str] = set()
    for source in sources:
        sheet = source.get("sheet")
        table = source.get("table")
        columns = source.get("columns")
        headers = source.get("headers")
        types = source.get("types")
        nullable = source.get("nullableColumns", [])
        if not isinstance(sheet, str) or not sheet:
            raise ValueError("Each source must have a worksheet name")
        if not isinstance(table, str) or not IDENTIFIER.fullmatch(table):
            raise ValueError(f"{sheet}: invalid target table")
        if table in tables or sheet in sheets:
            raise ValueError(f"{sheet}: duplicate source sheet or table")
        tables.add(table)
        sheets.add(sheet)
        if not all(isinstance(items, list) for items in (columns, headers, types)):
            raise ValueError(f"{sheet}: headers, columns and types must be arrays")
        if not columns or not len(columns) == len(headers) == len(types):
            raise ValueError(f"{sheet}: headers, columns and types must have equal length")
        if len(set(columns)) != len(columns) or not all(
            isinstance(column, str) and IDENTIFIER.fullmatch(column)
            for column in columns
        ):
            raise ValueError(f"{sheet}: target columns must be unique SQL identifiers")
        if not all(isinstance(header, str) and header for header in headers):
            raise ValueError(f"{sheet}: source headers must be non-empty strings")
        if len(set(headers)) != len(headers):
            raise ValueError(f"{sheet}: source headers must be unique")
        if not all(isinstance(item, str) for item in types):
            raise ValueError(f"{sheet}: source types must be strings")
        unknown_types = set(types) - SUPPORTED_TYPES
        if unknown_types:
            raise ValueError(f"{sheet}: unsupported types {sorted(unknown_types)}")
        if (
            not isinstance(nullable, list)
            or not all(isinstance(column, str) for column in nullable)
            or len(set(nullable)) != len(nullable)
            or not set(nullable).issubset(columns)
        ):
            raise ValueError(f"{sheet}: nullableColumns must reference target columns")
        primary = source.get("primaryColumn")
        if not isinstance(primary, int) or primary < 0 or primary >= len(columns):
            raise ValueError(f"{sheet}: invalid primaryColumn")
        if columns[primary] in nullable:
            raise ValueError(f"{sheet}: primary identifier cannot be nullable")
        if not isinstance(source.get("expectedRows"), int) or source["expectedRows"] <= 0:
            raise ValueError(f"{sheet}: expectedRows must be a positive integer")

    configured_total = sum(source["expectedRows"] for source in sources)
    if configured_total != expected_total:
        raise ValueError(
            f"Schema source rows total {configured_total}; expected {expected_total}"
        )


def _is_expected_type(value: Any, expected_type: str) -> bool:
    if expected_type == "string":
        return isinstance(value, str)
    if expected_type == "integer":
        return isinstance(value, int) and not isinstance(value, bool)
    if expected_type == "decimal":
        if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
            return False
        if isinstance(value, float):
            return math.isfinite(value)
        if isinstance(value, Decimal):
            return value.is_finite()
        return True
    if expected_type == "date":
        return isinstance(value, (date, datetime))
    if expected_type == "datetime":
        return isinstance(value, datetime)
    return False


def normalize_value(value: Any, expected_type: str, column: str) -> Any:
    if value is None:
        return None
    if expected_type == "date":
        return value.date() if isinstance(value, datetime) else value
    if expected_type == "datetime":
        return value.replace(tzinfo=None)
    if expected_type == "decimal":
        scale = 8 if column.endswith("_pct") else 4
        quantum = Decimal(1).scaleb(-scale)
        return Decimal(str(value)).quantize(quantum)
    return value


def validate_source_rows(
    source: dict[str, Any],
    headers: Sequence[Any],
    rows: Iterable[Sequence[Any]],
) -> list[tuple[Any, ...]]:
    expected_headers = tuple(source["headers"])
    actual_headers = tuple(headers)
    if actual_headers != expected_headers:
        raise ValueError(
            f"{source['sheet']} headers differ: got {actual_headers!r}; "
            f"expected {expected_headers!r}"
        )

    columns: list[str] = source["columns"]
    types: list[str] = source["types"]
    nullable = set(source.get("nullableColumns", []))
    primary_index = int(source["primaryColumn"])
    primary_values: set[Any] = set()
    validated: list[tuple[Any, ...]] = []

    for excel_row, row in enumerate(rows, start=2):
        if len(row) != len(columns):
            raise ValueError(
                f"{source['sheet']} row {excel_row} has {len(row)} cells; "
                f"expected {len(columns)}"
            )
        normalized: list[Any] = []
        for index, (value, expected_type) in enumerate(zip(row, types, strict=True)):
            column = columns[index]
            if value is None:
                if column not in nullable:
                    raise ValueError(
                        f"{source['sheet']} row {excel_row} column "
                        f"{source['headers'][index]} is required"
                    )
                normalized.append(None)
                continue
            if not _is_expected_type(value, expected_type):
                raise ValueError(
                    f"{source['sheet']} row {excel_row} column "
                    f"{source['headers'][index]} expected {expected_type}; "
                    f"got {type(value).__name__}"
                )
            if expected_type == "string" and not value.strip():
                raise ValueError(
                    f"{source['sheet']} row {excel_row} column "
                    f"{source['headers'][index]} cannot be blank"
                )
            normalized.append(normalize_value(value, expected_type, column))

        primary_value = normalized[primary_index]
        if isinstance(primary_value, str):
            if not primary_value.strip():
                raise ValueError(
                    f"{source['sheet']} row {excel_row} has an empty primary identifier"
                )
            uniqueness_key: Any = primary_value.rstrip().casefold()
        else:
            uniqueness_key = primary_value
        if uniqueness_key in primary_values:
            raise ValueError(
                f"{source['sheet']} row {excel_row} has duplicate primary identifier "
                f"{primary_value!r}"
            )
        primary_values.add(uniqueness_key)
        validated.append(tuple(normalized))

    if len(validated) != source["expectedRows"]:
        raise ValueError(
            f"{source['sheet']} has {len(validated)} rows; "
            f"expected {source['expectedRows']}"
        )
    return validated


def process_workbook(
    workbook_path: Path,
    schema: dict[str, Any],
    on_source: Callable[[dict[str, Any], list[tuple[Any, ...]]], None] | None = None,
) -> dict[str, int]:
    validate_schema(schema)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        required_sheets = {source["sheet"] for source in schema["sources"]}
        missing = sorted(required_sheets - set(workbook.sheetnames))
        if missing:
            raise ValueError(f"Missing source worksheets: {', '.join(missing)}")

        row_counts: dict[str, int] = {}
        for source in schema["sources"]:
            worksheet = workbook[source["sheet"]]
            iterator: Iterator[tuple[Any, ...]] = worksheet.iter_rows(values_only=True)
            headers = next(iterator, ())
            rows = validate_source_rows(source, headers, iterator)
            row_counts[source["sheet"]] = len(rows)
            if on_source:
                on_source(source, rows)
        total_rows = sum(row_counts.values())
        if total_rows != schema["expectedTotalRows"]:
            raise ValueError(
                f"Total row count is {total_rows}; expected {schema['expectedTotalRows']}"
            )
        return row_counts
    finally:
        workbook.close()


def batches(
    rows: Sequence[tuple[Any, ...]], size: int = BATCH_SIZE
) -> Iterator[Sequence[tuple[Any, ...]]]:
    for offset in range(0, len(rows), size):
        yield rows[offset : offset + size]


def insert_staging_source(
    connection: pyodbc.Connection,
    source: dict[str, Any],
    version_id: str,
    rows: Sequence[tuple[Any, ...]],
) -> None:
    columns = ["dataset_version", *source["columns"]]
    placeholders = ",".join("?" for _ in columns)
    quoted_columns = ",".join(f"[{column}]" for column in columns)
    statement = (
        f"INSERT INTO staging.[{source['table']}] ({quoted_columns}) "
        f"VALUES ({placeholders})"
    )
    cursor = connection.cursor()
    cursor.fast_executemany = True
    for row_batch in batches(rows):
        cursor.executemany(
            statement,
            [(version_id, *row) for row in row_batch],
        )


def register_loading_version(
    connection: pyodbc.Connection,
    version_id: str,
    source_blob_name: str,
    etag: str,
    sha256: str,
    size: int,
) -> None:
    connection.cursor().execute(
        """
        INSERT INTO app.dataset_version
          (version_id, blob_name, blob_etag, sha256, file_size_bytes, status)
        VALUES (?, ?, ?, ?, ?, N'loading')
        """,
        version_id,
        source_blob_name,
        etag,
        sha256,
        size,
    )
    connection.commit()


def mark_staged(
    connection: pyodbc.Connection,
    version_id: str,
    row_counts: dict[str, int],
) -> None:
    total_rows = sum(row_counts.values())
    cursor = connection.cursor()
    cursor.execute(
        """
        UPDATE app.dataset_version
        SET status = N'staged',
            total_rows = ?,
            source_row_counts = ?,
            validated_at = SYSUTCDATETIME()
        WHERE version_id = ? AND status = N'loading'
        """,
        total_rows,
        json.dumps(row_counts, ensure_ascii=False, sort_keys=True),
        version_id,
    )
    if cursor.rowcount != 1:
        raise RuntimeError("Dataset version was not in loading state")
    connection.commit()


def activate(
    connection: pyodbc.Connection,
    schema: dict[str, Any],
    version_id: str,
) -> None:
    cursor = connection.cursor()
    try:
        cursor.execute(
            "SET XACT_ABORT ON; "
            "SET TRANSACTION ISOLATION LEVEL SERIALIZABLE;"
        )
        lock_result = cursor.execute(
            """
            DECLARE @result int;
            EXEC @result = sys.sp_getapplock
              @Resource = N'norautron-dataset-activation',
              @LockMode = N'Exclusive',
              @LockOwner = N'Transaction',
              @LockTimeout = 60000;
            SELECT @result;
            """
        ).fetchone()
        if not lock_result or int(lock_result[0]) < 0:
            raise RuntimeError("Could not acquire the dataset activation lock")

        version = cursor.execute(
            """
            SELECT status, total_rows
            FROM app.dataset_version WITH (UPDLOCK, HOLDLOCK)
            WHERE version_id = ?
            """,
            version_id,
        ).fetchone()
        if not version or version[0] != "staged":
            raise RuntimeError("Dataset version was not in staged state")
        if int(version[1]) != schema["expectedTotalRows"]:
            raise RuntimeError("Staged dataset total does not match the schema")

        for source in schema["sources"]:
            table = source["table"]
            count = cursor.execute(
                f"SELECT COUNT_BIG(*) FROM staging.[{table}] "
                "WHERE dataset_version = ?",
                version_id,
            ).fetchone()[0]
            if int(count) != source["expectedRows"]:
                raise RuntimeError(
                    f"staging.{table} has {count} rows; "
                    f"expected {source['expectedRows']}"
                )
            columns = ["dataset_version", *source["columns"]]
            quoted = ",".join(f"[{column}]" for column in columns)
            cursor.execute(
                f"INSERT INTO raw.[{table}] ({quoted}) "
                f"SELECT {quoted} FROM staging.[{table}] "
                "WHERE dataset_version = ?",
                version_id,
            )
            if cursor.rowcount != source["expectedRows"]:
                raise RuntimeError(f"raw.{table} received an unexpected row count")

        cursor.execute(
            """
            UPDATE app.dataset_version
            SET status = N'archived'
            WHERE status = N'active'
            """
        )
        cursor.execute(
            """
            UPDATE app.dataset_version
            SET status = N'active',
                loaded_at = SYSUTCDATETIME(),
                error_message = NULL
            WHERE version_id = ? AND status = N'staged'
            """,
            version_id,
        )
        if cursor.rowcount != 1:
            raise RuntimeError("Dataset activation state transition failed")

        for source in schema["sources"]:
            cursor.execute(
                f"DELETE FROM staging.[{source['table']}] WHERE dataset_version = ?",
                version_id,
            )
        connection.commit()
    except Exception:
        connection.rollback()
        raise


def fail_version(connection: pyodbc.Connection, version_id: str, error: Exception) -> None:
    try:
        connection.rollback()
        connection.cursor().execute(
            """
            UPDATE app.dataset_version
            SET status = N'failed',
                error_message = ?
            WHERE version_id = ? AND status <> N'active'
            """,
            str(error)[:2000],
            version_id,
        )
        connection.commit()
    except Exception as status_error:
        connection.rollback()
        log_event(
            "dataset_failure_status_update_failed",
            level=logging.ERROR,
            version=version_id,
            error_type=type(status_error).__name__,
        )


def validate_local(workbook_path: Path) -> None:
    with observed_stage("local_validation", workbook=workbook_path.name):
        schema = load_schema()
        row_counts = process_workbook(workbook_path, schema)
    log_event(
        "workbook_validated",
        workbook=workbook_path.name,
        rows=sum(row_counts.values()),
        source_count=len(row_counts),
    )


def run_ingestion() -> None:
    run_id = os.getenv("CONTAINER_APP_JOB_EXECUTION_NAME") or str(uuid.uuid4())
    with observed_stage("schema_load", run_id=run_id):
        schema = load_schema()
    version_id = str(uuid.uuid4())
    url = required("DATA_BLOB_URL")
    source_blob_name = blob_name(url)
    work_dir = Path(os.getenv("INGEST_WORK_DIR", ".work"))
    workbook_path = work_dir / f"{version_id}.xlsx"
    connection: pyodbc.Connection | None = None

    log_event("ingestion_started", run_id=run_id, version=version_id)
    try:
        try:
            with observed_stage("blob_download", run_id=run_id, version=version_id):
                etag, sha256, size = download_workbook(workbook_path)
            log_event(
                "workbook_downloaded",
                run_id=run_id,
                version=version_id,
                blob=source_blob_name,
                etag=etag,
                sha256=sha256,
                bytes=size,
            )

            with observed_stage("sql_connect", run_id=run_id, version=version_id):
                connection = sql_connection()
            with observed_stage(
                "version_register", run_id=run_id, version=version_id
            ):
                register_loading_version(
                    connection, version_id, source_blob_name, etag, sha256, size
                )

            def stage_source(
                source: dict[str, Any], rows: list[tuple[Any, ...]]
            ) -> None:
                with observed_stage(
                    "source_stage",
                    run_id=run_id,
                    version=version_id,
                    source=source["sheet"],
                    rows=len(rows),
                ):
                    insert_staging_source(connection, source, version_id, rows)

            try:
                with observed_stage(
                    "workbook_process", run_id=run_id, version=version_id
                ):
                    row_counts = process_workbook(workbook_path, schema, stage_source)
                with observed_stage(
                    "version_stage", run_id=run_id, version=version_id
                ):
                    mark_staged(connection, version_id, row_counts)
                with observed_stage(
                    "dataset_activate", run_id=run_id, version=version_id
                ):
                    activate(connection, schema, version_id)
            except Exception as error:
                fail_version(connection, version_id, error)
                raise
        finally:
            if connection is not None:
                connection.close()
            workbook_path.unlink(missing_ok=True)
            try:
                work_dir.rmdir()
            except OSError:
                pass
    except Exception as error:
        if _run_count is not None:
            _run_count.add(1, {"outcome": "failure"})
        log_event(
            "ingestion_completed",
            level=logging.ERROR,
            run_id=run_id,
            version=version_id,
            outcome="failure",
            error_type=type(error).__name__,
        )
        raise

    if _run_count is not None:
        _run_count.add(1, {"outcome": "success"})
    log_event(
        "ingestion_completed",
        run_id=run_id,
        version=version_id,
        outcome="success",
        rows=schema["expectedTotalRows"],
        completed_at=datetime.now(timezone.utc).isoformat(),
    )


def main() -> None:
    try:
        if len(sys.argv) == 3 and sys.argv[1] == "--validate-only":
            validate_local(Path(sys.argv[2]))
            return
        if len(sys.argv) != 1:
            raise SystemExit("Usage: main.py [--validate-only <workbook.xlsx>]")
        run_ingestion()
    except Exception as error:
        log_event(
            "ingest_command_failed",
            level=logging.ERROR,
            error_type=type(error).__name__,
        )
        raise SystemExit(1) from None


if __name__ == "__main__":
    main()
